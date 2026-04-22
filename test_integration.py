"""네트워크 없이 샘플 데이터로 엑셀 + Slack blocks 구성 검증."""

import json
from datetime import date
from pathlib import Path

from scraper import IpoItem, classify_sections
from excel_writer import write_excel
import slack_notify


def make(name, start, end, fixed="-", desired="1,000", comp="", uw="X증권", no=1):
    return IpoItem(
        name=name,
        schedule=f"{start.replace('-','.')}~{end.split('-')[1]}.{end.split('-')[2]}",
        start_date=start, end_date=end,
        fixed_price=fixed, desired_price=desired, competition=comp,
        underwriter=uw,
        detail_url=f"http://www.38.co.kr/html/fund/?o=v&no={no}",
        analysis_url=f"http://www.38.co.kr/html/fund/index.htm?o=v&no={no}",
    )


today = date(2026, 4, 22)
items = [
    make("공모중A", "2026-04-20", "2026-04-22", fixed="15,000",
         desired="13,000~15,000", comp="", uw="NH투자증권", no=2287),
    make("공모중B", "2026-04-22", "2026-04-23", fixed="-",
         desired="7,000~8,000", comp="", uw="미래에셋증권", no=2286),
    make("예정A",   "2026-04-27", "2026-04-28", fixed="-",
         desired="5,300~6,000", comp="", uw="유진투자증권,NH투자증권", no=2278),
    make("마감A",   "2026-04-14", "2026-04-15", fixed="2,000",
         desired="2,000~2,000", comp="1727.59:1", uw="키움증권", no=2282),
    make("마감B",   "2026-04-15", "2026-04-16", fixed="16,600",
         desired="12,100~16,600", comp="1913:1", uw="NH투자증권,유진투자증권", no=2275),
    # 전체 엑셀용 과거 데이터
    make("과거1", "2025-06-01", "2025-06-02", fixed="10,000",
         desired="8,000~10,000", comp="1500:1", uw="KB증권", no=2200),
]

sections = classify_sections(items, today=today)
print("섹션 분류:")
for k, v in sections.items():
    print(f"  {k}: {[x.name for x in v]}")

# 엑셀 생성
out_dir = Path("/tmp/ipo_test")
out_dir.mkdir(exist_ok=True)
xlsx_path = out_dir / "20260422_38_ipo_schedule.xlsx"
write_excel(items, xlsx_path)
print(f"\n엑셀 생성: {xlsx_path} ({xlsx_path.stat().st_size} bytes)")

# openpyxl로 역검증
from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
ws = wb.active
print(f"  시트명: {ws.title}")
print(f"  헤더: {[ws.cell(1, c).value for c in range(1, 9)]}")
print(f"  데이터 행 수: {ws.max_row - 1}")
print(f"  첫 행: {[ws.cell(2, c).value for c in range(1, 9)]}")

# Slack blocks 구성
blocks = slack_notify.build_blocks(
    ongoing=sections["ongoing"],
    upcoming=sections["upcoming"],
    recent_end=sections["recent_end"],
    run_date="2026-04-22",
)
print(f"\nSlack blocks 수: {len(blocks)}")
for i, b in enumerate(blocks):
    t = b.get("type")
    if t == "header":
        print(f"  [{i}] header: {b['text']['text']}")
    elif t == "section":
        text = b["text"]["text"]
        preview = text[:80].replace("\n", " ↵ ")
        print(f"  [{i}] section: {preview}{'...' if len(text) > 80 else ''}")
    elif t == "divider":
        print(f"  [{i}] divider")
    elif t == "context":
        print(f"  [{i}] context: {b['elements'][0]['text'][:60]}...")

# Block Kit payload가 JSON으로 직렬화 되는지 확인 (Slack API가 받을 형태)
payload = json.dumps(blocks, ensure_ascii=False)
print(f"\nJSON 직렬화 크기: {len(payload)} bytes (Slack 상한 약 50KB, 여유 충분)")
assert len(payload) < 40000

# 스킵 조건 시뮬레이션
from main import _snapshot_payload
snap1 = _snapshot_payload(sections)
snap2 = _snapshot_payload(sections)
assert snap1 == snap2, "동일 입력 → 동일 스냅샷"
# 한 건만 바꿔보기
items[0] = make("공모중A", "2026-04-20", "2026-04-22", fixed="15,500",  # 가격 변경
                desired="13,000~15,000", comp="", uw="NH투자증권", no=2287)
sections2 = classify_sections(items, today=today)
snap3 = _snapshot_payload(sections2)
assert snap1 != snap3, "가격 변경 → 다른 스냅샷"
print("\n스냅샷 diff 로직 OK")

print("\n전체 통과 ✓")
