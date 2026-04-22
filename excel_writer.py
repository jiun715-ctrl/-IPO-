"""증권사별·연도별 IPO 주간 집계 엑셀 생성."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scraper import IpoItem


# 대상 연도: 배치일 기준 최근 3년(올해, 작년, 재작년)
TARGET_YEARS_COUNT = 3

# (헤더, 컬럼 너비)
COLUMNS = [
    ("해당연도", 10),
    ("증권사명", 24),
    ("총 주간횟수", 12),
    ("내역(청약시작일)", 80),
]

HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _split_underwriters(text: str) -> list[str]:
    """주간사 문자열을 개별 증권사 리스트로 분리.
    예: 'NH투자증권,미래에셋증권' → ['NH투자증권', '미래에셋증권']
    """
    if not text:
        return []
    # 쉼표/슬래시 기준 분리, 공백 제거, 빈값 제거
    parts = []
    for chunk in text.replace("/", ",").split(","):
        s = chunk.strip()
        if s:
            parts.append(s)
    return parts


def _format_start_md(start_date: str) -> str:
    """'2026-03-31' → '3/31' (한 자리 그대로)."""
    try:
        dt = datetime.strptime(start_date, "%Y-%m-%d")
        return f"{dt.month}/{dt.day}"
    except (ValueError, TypeError):
        return ""


def aggregate(items: Iterable[IpoItem], target_years: list[int]) -> list[dict]:
    """
    (연도, 증권사) 기준으로 집계.

    반환: [{"year": 2026, "underwriter": "NH투자증권", "count": 2,
            "items": [(start_date, name), ...]}, ...]

    - 한 종목에 증권사 N개가 있으면 각 증권사에 1건씩 계상
    - target_years에 없는 연도는 제외
    - 연도 desc → count desc → underwriter 가나다순
    - 내역은 청약시작일 asc
    """
    # (year, underwriter) → [(start_date_str, name), ...]
    bucket: dict[tuple[int, str], list[tuple[str, str]]] = defaultdict(list)

    for it in items:
        if not it.start_date:
            continue
        try:
            start_dt = datetime.strptime(it.start_date, "%Y-%m-%d")
        except ValueError:
            continue
        year = start_dt.year
        if year not in target_years:
            continue

        uw_list = _split_underwriters(it.underwriter)
        for uw in uw_list:
            bucket[(year, uw)].append((it.start_date, it.name))

    rows: list[dict] = []
    for (year, uw), entries in bucket.items():
        entries.sort(key=lambda x: x[0])  # 청약시작일 asc
        rows.append(
            {
                "year": year,
                "underwriter": uw,
                "count": len(entries),
                "items": entries,
            }
        )

    # 정렬: 연도 desc → count desc → 증권사 가나다순
    rows.sort(key=lambda r: (-r["year"], -r["count"], r["underwriter"]))
    return rows


def _format_items_cell(entries: list[tuple[str, str]]) -> str:
    """[('2026-03-31', '종목A'), ('2026-04-05', '종목B')] → '종목A(3/31), 종목B(4/5)'"""
    parts = [f"{name}({_format_start_md(sd)})" for sd, name in entries]
    return ", ".join(parts)


def write_excel(
    items: Iterable[IpoItem],
    out_path: Path,
    target_years: list[int],
) -> Path:
    """증권사별·연도별 집계를 xlsx로 저장."""
    items = list(items)
    rows = aggregate(items, target_years)

    wb = Workbook()
    ws = wb.active
    ws.title = "증권사별 IPO 집계"

    # 헤더
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["year"]).alignment = CENTER
        ws.cell(row=row_idx, column=2, value=row["underwriter"]).alignment = LEFT
        ws.cell(row=row_idx, column=3, value=f"{row['count']}건").alignment = CENTER
        ws.cell(row=row_idx, column=4, value=_format_items_cell(row["items"])).alignment = LEFT

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    return out_path
